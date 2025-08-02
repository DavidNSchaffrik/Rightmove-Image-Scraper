import re
import sys
import time
import shutil
from pathlib import Path
from typing import Optional, Tuple, List

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image


# ========== CONFIG ==========
ROOT = r"G:\My Drive\tiktok\rightmove\Rightmove-Image-Scraper\rightmove_images"
LINKS_CANDIDATES = ["Links.txt", "links.txt"]   # one URL per line; new links appended at the bottom
OUTPUT_XLSX = "rightmove_properties.xlsx"

IMAGE_FILENAME = "image_1.jpg"
INFO_FILENAME = "property_info.txt"

THUMB_MAX_W = 180   # embedded image max width (px)
THUMB_MAX_H = 120   # embedded image max height (px)

# Column layout
COL_WIDTHS = {
    "A": 8,     # Folder
    "B": 60,    # Link
    "C": 40,    # Location
    "D": 14,    # Price
    "E": 35,    # Image
}
TABLE_NAME = "RightmoveTable"
# ===========================


# ----- Utilities for thumbnails that persist until save -----
class TempThumbManager:
    def __init__(self, base_dir: Path):
        self.tmp_dir = base_dir / "__thumbs_tmp__"
        self.tmp_dir.mkdir(exist_ok=True)
        self.created: List[Path] = []

    def make_thumb(self, src: Path) -> Optional[Path]:
        try:
            with Image.open(src) as im:
                im_copy = im.copy()
                im_copy.thumbnail((THUMB_MAX_W, THUMB_MAX_H))
                # Ensure format compatibility (e.g., JPEG/PNG)
                ext = src.suffix.lower() if src.suffix.lower() in [".png", ".jpg", ".jpeg"] else ".png"
                out_path = self._unique_name(src.stem, ext)
                im_copy.save(out_path)
                self.created.append(out_path)
                return out_path
        except Exception as e:
            print(f"Warning: could not create thumbnail for {src}: {e}", file=sys.stderr)
            return None

    def _unique_name(self, stem: str, ext: str) -> Path:
        out = self.tmp_dir / f"{stem}__thumb{ext}"
        idx = 1
        while out.exists():
            out = self.tmp_dir / f"{stem}__thumb_{idx}{ext}"
            idx += 1
        return out

    def cleanup(self):
        for p in self.created:
            try:
                p.unlink(missing_ok=True)
            except Exception:
                pass
        try:
            if self.tmp_dir.exists() and not any(self.tmp_dir.iterdir()):
                self.tmp_dir.rmdir()
        except Exception:
            pass


# ----- Workbook helpers -----
def ensure_workbook(path: Path):
    """Open workbook if valid; otherwise create new. We won't fail if the old file is corrupt/locked."""
    if path.exists():
        try:
            wb = load_workbook(path)
            ws = wb.active
            _ensure_headers(ws)
            ws.freeze_panes = "A2"
            _apply_pretty_defaults(ws)
            return wb, ws
        except Exception as e:
            print(f"Workbook at {path} is unreadable ({e}). Will create a fresh file on save.")
    # New workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Rightmove"
    _ensure_headers(ws)
    ws.freeze_panes = "A2"
    _apply_pretty_defaults(ws)
    return wb, ws


def _ensure_headers(ws):
    if ws.max_row < 1 or (ws.cell(row=1, column=1).value or "") != "Folder":
        ws.delete_rows(1, ws.max_row)  # clear if anything odd
        ws.append(["Folder", "Link", "Location", "Price", "Image"])
    # Header styling
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFECECEC", end_color="FFECECEC", fill_type="solid")
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill


def _apply_pretty_defaults(ws):
    # Column widths
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width
    # Wrap and top-align location to look nice next to image
    for r in range(2, ws.max_row + 1):
        c_loc = ws.cell(row=r, column=3)
        c_loc.alignment = Alignment(wrap_text=True, vertical="top")


def collect_existing_links(ws):
    existing = set()
    for r in range(2, ws.max_row + 1):
        link = ws.cell(row=r, column=2).value
        if link:
            existing.add(str(link).strip())
    return existing


def read_links_with_line_numbers(root: Path):
    links_path = None
    for name in LINKS_CANDIDATES:
        p = root / name
        if p.exists():
            links_path = p
            break
    if not links_path:
        print(f"Links file not found. Tried: {', '.join(LINKS_CANDIDATES)} in {root}", file=sys.stderr)
        return []
    out = []
    with links_path.open("r", encoding="utf-8", errors="ignore") as f:
        for idx, raw in enumerate(f, start=1):
            out.append((idx, raw.rstrip("\n")))
    return out


def parse_property_info(info_path: Path) -> Tuple[Optional[str], Optional[str]]:
    if not info_path.exists():
        return None, None
    location = None
    price_text = None
    with info_path.open("r", encoding="utf-8", errors="ignore") as f:
        for raw in f:
            line = raw.strip()
            low = line.lower()
            if low.startswith("address:"):
                location = line.split(":", 1)[1].strip()
            elif low.startswith("price:"):
                price_text = line.split(":", 1)[1].strip()
    return location, price_text


def parse_price_to_number(price_text: Optional[str]) -> Optional[float]:
    if not price_text:
        return None
    cleaned = re.sub(r"[^\d.]", "", price_text)
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def add_thumbnail(ws, img_path: Path, row: int, col_letter: str, thumbs: TempThumbManager):
    if not img_path.exists():
        return
    thumb_path = thumbs.make_thumb(img_path)
    if not thumb_path:
        return
    try:
        xl_img = XLImage(str(thumb_path))
        ws.add_image(xl_img, f"{col_letter}{row}")
    except Exception as e:
        print(f"Warning: could not embed image {img_path}: {e}", file=sys.stderr)


def set_row_height_for_image(ws, row: int):
    """
    Excel row height is in points (1 pt = 1/72 in). Typical screen 96 dpi.
    px -> pt ≈ px * (72/96) = px * 0.75. Add padding for text/borders.
    """
    desired = int(THUMB_MAX_H * 0.75) + 10  # padding
    if ws.row_dimensions[row].height is None or ws.row_dimensions[row].height < desired:
        ws.row_dimensions[row].height = desired


def update_table(ws):
    # Create/update an Excel table for nicer styling with banded rows
    ref = f"A1:E{ws.max_row}"
    if TABLE_NAME in ws.tables:
        tbl = ws.tables[TABLE_NAME]
        tbl.ref = ref
    else:
        tbl = Table(displayName=TABLE_NAME, ref=ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tbl.tableStyleInfo = style
        ws.add_table(tbl)


def try_backup(src: Path) -> Optional[Path]:
    try:
        backup = src.with_suffix(src.suffix + ".bak")
        if backup.exists():
            backup.unlink()
        shutil.copy2(src, backup)
        print(f"Backed up to {backup}")
        return backup
    except Exception as e:
        print(f"Could not backup: {e}")
        return None


def try_save_workbook(wb: Workbook, out_path: Path) -> Path:
    try:
        wb.save(out_path)
        return out_path
    except Exception as e:
        print(f"Primary save failed for {out_path}: {e}")
    if out_path.exists():
        try_backup(out_path)
    alt = out_path.with_name(out_path.stem + "_new.xlsx")
    try:
        wb.save(alt)
        print(f"Saved to {alt} instead. Close any open copies of '{out_path.name}' and rename if desired.")
        return alt
    except Exception:
        ts = time.strftime("%Y%m%d_%H%M%S")
        alt_ts = out_path.with_name(f"{out_path.stem}_{ts}.xlsx")
        wb.save(alt_ts)
        print(f"Saved to {alt_ts} instead. Close any open copies of '{out_path.name}' and rename if desired.")
        return alt_ts


def main():
    root = Path(ROOT)
    if not root.exists():
        print(f"Root path not found: {root}", file=sys.stderr)
        sys.exit(1)

    out_path = root / OUTPUT_XLSX
    wb, ws = ensure_workbook(out_path)
    existing_links = collect_existing_links(ws)

    lines = read_links_with_line_numbers(root)
    if not lines:
        print("No links to process.")
        sys.exit(0)

    thumbs = TempThumbManager(root)
    rows_added = 0

    try:
        for line_no, raw in lines:
            link = (raw or "").strip()
            folder = root / str(line_no)  # line 1 -> folder "1", etc.

            if not link:
                continue
            if link in existing_links:
                continue
            if not folder.exists():
                print(f"Warning: missing folder for line {line_no}: {folder}", file=sys.stderr)
                continue

            info_path = folder / INFO_FILENAME
            img_path = folder / IMAGE_FILENAME

            location, price_text = parse_property_info(info_path)
            price_num = parse_price_to_number(price_text)

            next_row = ws.max_row + 1
            # Folder
            ws.cell(row=next_row, column=1, value=str(line_no))

            # Link (clickable)
            c_link = ws.cell(row=next_row, column=2, value=link)
            if link.lower().startswith("http"):
                c_link.hyperlink = link
                c_link.style = "Hyperlink"

            # Location (wrapped, top-aligned)
            c_loc = ws.cell(row=next_row, column=3, value=location or "")
            c_loc.alignment = Alignment(wrap_text=True, vertical="top")

            # Price (numeric if parsed, else raw)
            price_cell_value = price_num if price_num is not None else (price_text or "")
            c_price = ws.cell(row=next_row, column=4, value=price_cell_value)
            if price_num is not None:
                c_price.number_format = u'£#,##0'

            # Image
            add_thumbnail(ws, img_path, next_row, "E", thumbs)
            set_row_height_for_image(ws, next_row)

            existing_links.add(link)
            rows_added += 1

        # Freeze panes, widths, table and autofilter
        ws.freeze_panes = "A2"
        for col, width in COL_WIDTHS.items():
            ws.column_dimensions[col].width = width
        ws.auto_filter.ref = f"A1:E{ws.max_row}"
        update_table(ws)

        actual_path = try_save_workbook(wb, out_path)
        print(f"Updated: {actual_path} | Added {rows_added} new row(s).")

    finally:
        thumbs.cleanup()


if __name__ == "__main__":
    main()
